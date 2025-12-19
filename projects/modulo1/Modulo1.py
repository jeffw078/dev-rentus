import logging
from pathlib import Path
from datetime import date, timedelta, datetime
from typing import Optional, List

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font

PROCESS_LOGS: List[str] = []


def log(msg: str):
    print(msg)
    PROCESS_LOGS.append(msg)


# =============================================================================
# LEITURA ROBUSTA DE EXCEL (XLS / XLSX / EXTENSÃO ERRADA)
# =============================================================================
def read_excel_safe(path: Path) -> pd.DataFrame:
    log(f"[INFO] Lendo arquivo: {path.name}")
    try:
        # tentativa padrão (openpyxl / engine automático)
        return pd.read_excel(path)
    except Exception:
        try:
            # fallback para xls antigo real
            return pd.read_excel(path, engine="xlrd")
        except Exception as e:
            raise ValueError(f"Erro ao ler {path}: {e}")


def _find_column(df: pd.DataFrame, candidates):
    if isinstance(candidates, str):
        candidates = [candidates]

    colmap = {str(c).strip().lower(): c for c in df.columns}

    for cand in candidates:
        key = cand.lower()
        if key in colmap:
            return colmap[key]

    raise KeyError(f"Colunas não encontradas: {candidates}")


def _previous_month_range(today: Optional[date] = None):
    if today is None:
        today = date.today()

    first_this = date(today.year, today.month, 1)
    last_prev = first_this - timedelta(days=1)
    first_prev = date(last_prev.year, last_prev.month, 1)

    return first_prev, last_prev


# =============================================================================
# PROCESSAMENTO DO MÓDULO 1
# =============================================================================
def process_modulo1(
    ops_path: Path,
    hk_avulso_path: Path,
    demitidos_path: Path,
    aviso_previo_path: Path,
    situacao_path: Path,
    fp_path: Path,
    output_dir: Path,
):

    PROCESS_LOGS.clear()
    log("[INFO] Iniciando módulo 1")

    output_dir.mkdir(parents=True, exist_ok=True)

    # -------------------------------------------------------------------------
    # ETAPA 1 – OPS
    # -------------------------------------------------------------------------
    ops_df = read_excel_safe(ops_path)

    col_re = _find_column(ops_df, "codvigil")
    col_nome = _find_column(ops_df, "rsocial")
    col_valpag = _find_column(ops_df, "valpag")

    ops_df[col_re] = ops_df[col_re].astype(str).str.strip()

    resumo = (
        ops_df.groupby(col_re)
        .agg({col_nome: "first", col_valpag: "sum"})
        .reset_index()
    )

    final_df = pd.DataFrame({
        "RE": resumo[col_re],
        "NOME": resumo[col_nome],
        "Valor OP": resumo[col_valpag],
        "Valor Benefício": "",
        "Diferença": "",
        "Resultado": "",
        "Alerta": "",
        "DEMITIDO?": "ATIVO",
        "AVISO PRÉVIO?": "ATIVO",
        "SITUAÇÃO": "",
        "Cargo": "",
        "Escala": "",
        "Qtd. Dias Trabalhados": 0,
        "Falta Injustificada": 0,
        "Falta Justificada": 0,
        "Falta Abonada": 0,
        "INSS": 0,
        "Suspensão": 0,
        "Folga Trabalhada": 0,
        "Folga Trabalhada Cash": 0,
    })

    final_df.set_index("RE", inplace=True)

    # -------------------------------------------------------------------------
    # ETAPA 2 – hk_avulso
    # -------------------------------------------------------------------------
    hk_df = read_excel_safe(hk_avulso_path)

    hk_re = _find_column(hk_df, ["re", "codvigil"])
    hk_dt = _find_column(hk_df, ["dtavus", "data", "dt_avus"])
    hk_val = _find_column(hk_df, ["valor_total", "valor"])

    hk_df[hk_re] = hk_df[hk_re].astype(str).str.strip()
    hk_df[hk_val] = pd.to_numeric(hk_df[hk_val], errors="coerce")
    hk_df[hk_dt] = pd.to_datetime(hk_df[hk_dt], errors="coerce")

    hoje = date.today()

    for re in final_df.index:
        valor_op = final_df.at[re, "Valor OP"]
        dados_re = hk_df[hk_df[hk_re] == str(re)]

        if dados_re.empty:
            final_df.at[re, "Alerta"] = "nada encontrado"
            continue

        soma_hoje = dados_re[dados_re[hk_dt].dt.date == hoje][hk_val].sum()

        if soma_hoje == valor_op and soma_hoje != 0:
            final_df.at[re, "Valor Benefício"] = soma_hoje
            final_df.at[re, "Diferença"] = 0
            final_df.at[re, "Resultado"] = "OK"
            final_df.at[re, "Alerta"] = ""
            continue

        dados_validos = dados_re.dropna(subset=[hk_val, hk_dt]).copy()
        dados_validos["dia"] = dados_validos[hk_dt].dt.date
        soma_por_dia = dados_validos.groupby("dia")[hk_val].sum()

        iguais = soma_por_dia[soma_por_dia == valor_op]
        if not iguais.empty:
            dia = iguais.index[0].strftime("%d/%m/%Y")
            final_df.at[re, "Resultado"] = f"o valor foi lançado no dia {dia}"
            final_df.at[re, "Diferença"] = 0
            continue

        if len(soma_por_dia) == 1:
            soma_unica = soma_por_dia.iloc[0]
            dia = soma_por_dia.index[0].strftime("%d/%m/%Y")
            final_df.at[re, "Alerta"] = f"o valor {soma_unica:.2f} foi localizado na data {dia}"
            continue

        if len(soma_por_dia) > 1:
            soma_total = soma_por_dia.sum()
            final_df.at[re, "Alerta"] = f"o valor {soma_total:.2f} foi localizado em diversas datas"
            continue

        final_df.at[re, "Alerta"] = "nada encontrado"

    # -------------------------------------------------------------------------
    # ETAPA 3 – Diferença
    # -------------------------------------------------------------------------
    mask = final_df["Valor Benefício"] != ""
    final_df.loc[mask, "Diferença"] = (
        final_df.loc[mask, "Valor OP"]
        - pd.to_numeric(final_df.loc[mask, "Valor Benefício"], errors="coerce")
    )

    # -------------------------------------------------------------------------
    # ETAPA 4 – DEMITIDOS
    # -------------------------------------------------------------------------
    dem_df = read_excel_safe(demitidos_path)
    dem_re = _find_column(dem_df, ["re", "codvigil", "RE"])
    dem_df[dem_re] = dem_df[dem_re].astype(str).str.strip()
    dem_set = set(dem_df[dem_re])

    for re in final_df.index:
        if str(re) in dem_set:
            final_df.at[re, "DEMITIDO?"] = "DEMITIDO"

    # -------------------------------------------------------------------------
    # ETAPA 5 – AVISO PRÉVIO
    # -------------------------------------------------------------------------
    aviso_df = read_excel_safe(aviso_previo_path)
    aviso_re = _find_column(aviso_df, ["re", "codvigil", "RE"])
    aviso_fim = _find_column(aviso_df, ["data_fim", "dtfim", "fim"])

    aviso_df[aviso_re] = aviso_df[aviso_re].astype(str).str.strip()
    aviso_df[aviso_fim] = pd.to_datetime(aviso_df[aviso_fim], errors="coerce")

    ultimos = aviso_df.groupby(aviso_re)[aviso_fim].max()

    for re in final_df.index:
        if str(re) in ultimos and pd.notna(ultimos[str(re)]):
            final_df.at[re, "AVISO PRÉVIO?"] = ultimos[str(re)].strftime("%d/%m/%Y")

    # -------------------------------------------------------------------------
    # ETAPA 6 – SITUAÇÃO
    # -------------------------------------------------------------------------
    sit_df = read_excel_safe(situacao_path)

    sit_re = _find_column(sit_df, ["re", "codvigil", "RE"])
    sit_desc = _find_column(sit_df, ["descsituacao", "situacao", "situação"])
    sit_cargo = _find_column(sit_df, ["nomecargo", "cargo"])
    sit_escala = _find_column(sit_df, ["nomeescala", "escala"])

    sit_df[sit_re] = sit_df[sit_re].astype(str).str.strip()
    last_rows = sit_df.groupby(sit_re).tail(1).set_index(sit_re)

    for re in final_df.index:
        if str(re) in last_rows.index:
            linha = last_rows.loc[str(re)]
            final_df.at[re, "SITUAÇÃO"] = linha.get(sit_desc, "")
            final_df.at[re, "Cargo"] = linha.get(sit_cargo, "")
            final_df.at[re, "Escala"] = linha.get(sit_escala, "")

    # -------------------------------------------------------------------------
    # ETAPA 7 – FREQUÊNCIA (fp)
    # -------------------------------------------------------------------------
    fp_df = read_excel_safe(fp_path)

    fp_re = _find_column(fp_df, ["re", "codvigil", "RE"])
    fp_sit = _find_column(fp_df, ["situacao", "situação", "SITUACAO"])
    fp_sithoje = _find_column(fp_df, ["sithoje", "sit_hoje", "SITHOJE"])
    fp_data = _find_column(fp_df, ["data", "DATA"])

    fp_df[fp_re] = fp_df[fp_re].astype(str).str.strip()
    fp_df[fp_sit] = fp_df[fp_sit].astype(str).str.upper().str.strip()
    fp_df[fp_sithoje] = fp_df[fp_sithoje].astype(str).str.upper().str.strip()

    raw_dates = fp_df[fp_data]
    fp_df[fp_data] = pd.to_datetime(raw_dates, dayfirst=True, errors="coerce")

    inicio, fim = _previous_month_range()
    periodo_df = fp_df[
        (fp_df[fp_data] >= pd.Timestamp(inicio)) &
        (fp_df[fp_data] <= pd.Timestamp(fim))
    ]

    grouped_fp = periodo_df.groupby(fp_re)

    for re in final_df.index:
        if str(re) not in grouped_fp.groups:
            continue

        g = grouped_fp.get_group(str(re))
        final_df.at[re, "Qtd. Dias Trabalhados"] = len(g)

    # -------------------------------------------------------------------------
    # SALVAR ARQUIVO FINAL
    # -------------------------------------------------------------------------
    final_df.reset_index(inplace=True)

    hoje_str = date.today().strftime("%d-%m-%y")
    out_path = output_dir / f"analise-{hoje_str}-vr.xlsx"
    final_df.to_excel(out_path, index=False)

    wb = load_workbook(out_path)
    ws = wb.active

    header_fill = PatternFill("solid", fgColor="333333")
    header_font = Font(color="FFFFFF", bold=True)

    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font

    wb.save(out_path)

    log(f"[INFO] Arquivo final salvo em: {out_path}")
    log("[INFO] PROCESSO FINALIZADO COM SUCESSO!")

    return out_path, PROCESS_LOGS
