import os
from pathlib import Path

import pandas as pd
import unidecode
from openpyxl.styles import Font, PatternFill, Alignment, Border
from openpyxl.utils import get_column_letter


# =========================================================
# UTILIDADES
# =========================================================

def norm(s):
    """Normaliza texto para comparações (sem acento, maiúsculo, sem espaços extras)."""
    return unidecode.unidecode(str(s)).upper().strip()


# =========================================================
# VERSÃO 1 – ADAPTAÇÃO DIRETA DO SEU SCRIPT (RECOMENDADA)
# =========================================================

def processar_ficha_presenca_v1(input_path, output_path):
    """
    Versão 1:
    - Lê o arquivo FP (xlsx/xls) no caminho input_path.
    - Usa exatamente a lógica do seu script FP_resultado_clientes.
    - Gera um XLSX consolidado em output_path, com formatação visual.
    """

    # Garantir que são Path
    input_path = Path(input_path)
    output_path = Path(output_path)

    # -----------------------------------------
    # 1. Ler o arquivo FP (XLSX ou XLS)
    # -----------------------------------------
    ext = input_path.suffix.lower()

    if ext == ".xlsx":
        print(f"Lendo arquivo {input_path.name} (formato novo)...")
        df = pd.read_excel(input_path, engine="openpyxl")
    elif ext == ".xls":
        print(f"Lendo arquivo {input_path.name} (formato antigo)...")
        df = pd.read_excel(input_path, engine="xlrd")
    else:
        raise ValueError("Formato de arquivo não suportado. Use .xlsx ou .xls para FP.")

    # -----------------------------------------
    # 2. Renomear as colunas de interesse
    #
    # A (0)  -> RE
    # B (1)  -> NOME
    # D (3)  -> DESC_CARGO
    # M (12) -> CLIENTE
    # Q (16) -> SITUACAO
    # S (18) -> SITHOJE
    # AD(29) -> NOMEESCAL
    # -----------------------------------------

    if len(df.columns) <= 29:
        raise ValueError("O arquivo não possui ao menos 30 colunas (até AD). Verifique o layout.")

    df = df.rename(columns={
        df.columns[0]:  "re",         # A
        df.columns[1]:  "nome",       # B
        df.columns[3]:  "desc_cargo", # D - descrição do cargo
        df.columns[12]: "cliente",    # M
        df.columns[16]: "situacao",   # Q
        df.columns[18]: "sithoje",    # S
        df.columns[29]: "escala",     # AD
    })

    # -----------------------------------------
    # Limpeza / Normalização básica
    # -----------------------------------------

    # RE e NOME
    df["re"]   = df["re"].astype(str).str.strip()
    df["nome"] = df["nome"].astype(str).str.upper().str.strip()

    # Demais campos textuais em MAIÚSCULO
    for col in ["cliente", "desc_cargo", "escala", "situacao", "sithoje"]:
        df[col] = df[col].astype(str).str.upper().str.strip()

    # Remover linhas sem RE ou sem cliente (opcional)
    df = df[df["re"] != ""]
    df = df[df["cliente"] != ""]

    # -----------------------------------------
    # 3. Mapa RE -> NOME (primeiro nome encontrado para aquele RE)
    # -----------------------------------------

    map_re_para_nome = (
        df.groupby("re", sort=False)["nome"]
          .first()
          .to_dict()
    )

    # -----------------------------------------
    # 4. Cliente mais frequente por RE (e segundo cliente, se empatar)
    # -----------------------------------------

    contagens_clientes = (
        df.groupby(["re", "cliente"])
          .size()
          .reset_index(name="qtd")
    )

    contagens_clientes_ordenadas = contagens_clientes.sort_values(
        ["re", "qtd"], ascending=[True, False]
    )

    top2_clientes_por_re = (
        contagens_clientes_ordenadas
        .groupby("re", group_keys=False)
        .head(2)   # no máximo 2 clientes por RE
    )

    # -----------------------------------------
    # 5. ESCALA mais frequente por RE
    # -----------------------------------------

    contagens_escala = (
        df.groupby(["re", "escala"])
          .size()
          .reset_index(name="qtd")
    )

    contagens_escala_ordenadas = contagens_escala.sort_values(
        ["re", "qtd"], ascending=[True, False]
    )

    top_escala_por_re = (
        contagens_escala_ordenadas
        .groupby("re", group_keys=False)
        .head(1)
    )

    escala_por_re = top_escala_por_re.set_index("re")["escala"].to_dict()

    # -----------------------------------------
    # 6. CARGO mais frequente por RE
    # -----------------------------------------

    contagens_cargo = (
        df.groupby(["re", "desc_cargo"])
          .size()
          .reset_index(name="qtd")
    )

    contagens_cargo_ordenadas = contagens_cargo.sort_values(
        ["re", "qtd"], ascending=[True, False]
    )

    top_cargo_por_re = (
        contagens_cargo_ordenadas
        .groupby("re", group_keys=False)
        .head(1)
    )

    cargo_por_re = top_cargo_por_re.set_index("re")["desc_cargo"].to_dict()

    # -----------------------------------------
    # 7. Normalizar SITUACAO e SITHOJE para contar INSS, Férias, etc.
    # -----------------------------------------

    df["situacao_norm"] = df["situacao"].apply(norm)
    df["sithoje_norm"]  = df["sithoje"].apply(norm)

    # Versão bruta em maiúsculas, preservando caracteres estranhos (F╔RIAS, SUSPENS├O)
    situ_up = df["situacao"].astype(str).str.upper()

    # INSS
    df["flag_inss"] = situ_up.str.contains("INSS", na=False, regex=False)

    # Férias: FÉRIAS / FERIAS / F╔RIAS
    df["flag_ferias"] = (
        situ_up.str.contains("FERIAS", na=False, regex=False) |   # sem acento
        situ_up.str.contains("FÉRIAS", na=False, regex=False) |   # com acento
        situ_up.str.contains("F╔RIAS", na=False, regex=False)     # bug de encoding
    )

    # Suspensão: SUSPENSÃO / SUSPENSAO / SUSPENS├O
    df["flag_suspensao"] = (
        situ_up.str.contains("SUSPENSAO", na=False, regex=False) |
        situ_up.str.contains("SUSPENSÃO", na=False, regex=False) |
        situ_up.str.contains("SUSPENS├O", na=False, regex=False)
    )

    # Combinações Q + S

    df["flag_falta_injustificada"] = (
        (df["situacao_norm"] == "FALTA") &
        (df["sithoje_norm"] == "TRABALHO")
    )

    df["flag_falta_abonada"] = (
        (df["situacao_norm"] == "FALTA") &
        (df["sithoje_norm"] == "FALTA ABONADA")
    )

    df["flag_falta_justificada"] = (
        (df["situacao_norm"] == "FALTA") &
        (df["sithoje_norm"].isin(["FALTA JUSTIFICADO", "FALTA JUSTIFICADA"]))
    )

    # Agrupar contagens por RE
    inss_por_re         = df.groupby("re")["flag_inss"].sum().to_dict()
    ferias_por_re       = df.groupby("re")["flag_ferias"].sum().to_dict()
    suspensao_por_re    = df.groupby("re")["flag_suspensao"].sum().to_dict()
    falta_injust_por_re = df.groupby("re")["flag_falta_injustificada"].sum().to_dict()
    falta_abon_por_re   = df.groupby("re")["flag_falta_abonada"].sum().to_dict()
    falta_just_por_re   = df.groupby("re")["flag_falta_justificada"].sum().to_dict()

    # -----------------------------------------
    # 8. Montar a saída consolidada por RE
    # -----------------------------------------

    linhas_resultado = []

    for re_val, grupo_re in top2_clientes_por_re.groupby("re", sort=False):
        grupo_re = grupo_re.sort_values("qtd", ascending=False)

        # Cliente mais frequente
        top = grupo_re.iloc[0]
        cliente_top = top["cliente"]
        qtd_top = top["qtd"]

        cliente_segundo = ""

        # Se existir um segundo cliente, verificar empate de quantidade
        if len(grupo_re) > 1:
            segundo = grupo_re.iloc[1]
            if segundo["qtd"] == qtd_top:
                cliente_segundo = segundo["cliente"]

        linhas_resultado.append({
            "re": re_val,
            "nome": map_re_para_nome.get(re_val, ""),
            "cliente_mais_frequente": cliente_top,
            "cliente_segundo_mais_frequente": cliente_segundo,
            "ESCALA": escala_por_re.get(re_val, ""),
            "CARGO": cargo_por_re.get(re_val, ""),
            "INSS": int(inss_por_re.get(re_val, 0)),
            "Férias": int(ferias_por_re.get(re_val, 0)),
            "Suspensão": int(suspensao_por_re.get(re_val, 0)),
            "Falta Injustificada": int(falta_injust_por_re.get(re_val, 0)),
            "Falta Abonada": int(falta_abon_por_re.get(re_val, 0)),
            "Falta Justificada": int(falta_just_por_re.get(re_val, 0)),
        })

    resultado = pd.DataFrame(linhas_resultado)

    # -----------------------------------------
    # 9. Exportar resultado para XLSX com formatação
    # -----------------------------------------

    output_path.parent.mkdir(parents=True, exist_ok=True)

    from openpyxl import Workbook
    from openpyxl.styles import Side

    # Criar workbook manualmente para ter controle total da formatação
    wb = Workbook()
    ws = wb.active
    ws.title = "Resultado"

    # Escrever cabeçalho
    headers = list(resultado.columns)
    ws.append(headers)

    # Escrever dados
    for _, row in resultado.iterrows():
        ws.append(list(row.values))

    # Fonte Calibri 8 e sem borda para tudo
    empty_border = Border()
    for row in ws.iter_rows():
        for cell in row:
            cell.font = Font(name="Calibri", size=8)
            cell.border = empty_border  # remove qualquer borda existente

    # Cabeçalho azul claro, negrito, centralizado, sem borda
    header_fill = PatternFill(fill_type="solid", fgColor="8EA9DB")  # Azul claro
    for cell in ws[1]:
        cell.font = Font(name="Calibri", size=8, bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = empty_border  # garante sem borda

    # Remover linhas de grade
    ws.sheet_view.showGridLines = False

    # Ajuste automático da largura das colunas (simulando ALT C O T)
    for col_idx, column_cells in enumerate(ws.columns, start=1):
        max_len = 0
        for cell in column_cells:
            value = cell.value
            if value is None:
                continue
            length = len(str(value))
            if length > max_len:
                max_len = length
        adjusted_width = max_len + 4  # margem extra
        ws.column_dimensions[get_column_letter(col_idx)].width = adjusted_width

    # Ajuste da altura das linhas (simulando ALT C O A com altura fixa)
    for row_idx in range(1, ws.max_row + 1):
        ws.row_dimensions[row_idx].height = 10.2

    wb.save(output_path)

    print()
    print(f'Arquivo "{output_path.name}" gerado com sucesso!')
    print(f"Total de linhas no arquivo de entrada: {len(df):,}")
    print(f"Total de REs distintos no resultado:   {len(resultado):,}")

    return output_path


# =========================================================
# VERSÃO 2 – MESMA LÓGICA, MAS TENTANDO USAR NOMES DE COLUNA
# =========================================================

def processar_ficha_presenca_v2(input_path, output_path):
    """
    Versão 2:
    - Tenta primeiro mapear colunas por NOME (RE, NOME, DESC_CARGO, CLIENTE, SITUACAO, SITHOJE, NOMEESCAL),
      se não achar, cai no mapeamento por posição (como na V1).
    - Gera saída com a mesma formatação da V1.
    """

    input_path = Path(input_path)
    output_path = Path(output_path)

    ext = input_path.suffix.lower()
    if ext == ".xlsx":
        df = pd.read_excel(input_path, engine="openpyxl")
    elif ext == ".xls":
        df = pd.read_excel(input_path, engine="xlrd")
    else:
        raise ValueError("Formato de arquivo não suportado. Use .xlsx ou .xls para FP.")

    original_cols = [c.strip().upper() for c in df.columns]
    df.columns = original_cols

    # Tentamos mapear por nome primeiro
    nome_map = {}

    def tem(col_name):
        return col_name in df.columns

    if tem("RE") and tem("NOME") and tem("CLIENTE") and tem("SITUACAO") and tem("SITHOJE") and (tem("NOMEESCAL") or tem("NOMEESCALA")):
        # Usar nomes
        nome_map["re"]   = "RE"
        nome_map["nome"] = "NOME"

        if tem("DESC_CARGO"):
            nome_map["desc_cargo"] = "DESC_CARGO"
        elif tem("CARGO"):
            nome_map["desc_cargo"] = "CARGO"

        nome_map["cliente"] = "CLIENTE"
        nome_map["situacao"] = "SITUACAO"
        nome_map["sithoje"]  = "SITHOJE"

        if tem("NOMEESCAL"):
            nome_map["escala"] = "NOMEESCAL"
        else:
            nome_map["escala"] = "NOMEESCALA"

        df = df.rename(columns={v: k for k, v in nome_map.items()})

    else:
        # Cai na mesma lógica da V1 (por posição)
        if len(df.columns) <= 29:
            raise ValueError("O arquivo não possui ao menos 30 colunas (até AD). Verifique o layout.")

        df = df.rename(columns={
            df.columns[0]:  "re",
            df.columns[1]:  "nome",
            df.columns[3]:  "desc_cargo",
            df.columns[12]: "cliente",
            df.columns[16]: "situacao",
            df.columns[18]: "sithoje",
            df.columns[29]: "escala",
        })

    # Limpeza
    df["re"]   = df["re"].astype(str).str.strip()
    df["nome"] = df["nome"].astype(str).str.upper().str.strip()

    for col in ["cliente", "desc_cargo", "escala", "situacao", "sithoje"]:
        df[col] = df[col].astype(str).str.upper().str.strip()

    df = df[df["re"] != ""]
    df = df[df["cliente"] != ""]

    map_re_para_nome = (
        df.groupby("re", sort=False)["nome"]
          .first()
          .to_dict()
    )

    contagens_clientes = (
        df.groupby(["re", "cliente"])
          .size()
          .reset_index(name="qtd")
    )

    contagens_clientes_ordenadas = contagens_clientes.sort_values(
        ["re", "qtd"], ascending=[True, False]
    )

    top2_clientes_por_re = (
        contagens_clientes_ordenadas
        .groupby("re", group_keys=False)
        .head(2)
    )

    contagens_escala = (
        df.groupby(["re", "escala"])
          .size()
          .reset_index(name="qtd")
    )

    contagens_escala_ordenadas = contagens_escala.sort_values(
        ["re", "qtd"], ascending=[True, False]
    )

    top_escala_por_re = (
        contagens_escala_ordenadas
        .groupby("re", group_keys=False)
        .head(1)
    )

    escala_por_re = top_escala_por_re.set_index("re")["escala"].to_dict()

    contagens_cargo = (
        df.groupby(["re", "desc_cargo"])
          .size()
          .reset_index(name="qtd")
    )

    contagens_cargo_ordenadas = contagens_cargo.sort_values(
        ["re", "qtd"], ascending=[True, False]
    )

    top_cargo_por_re = (
        contagens_cargo_ordenadas
        .groupby("re", group_keys=False)
        .head(1)
    )

    cargo_por_re = top_cargo_por_re.set_index("re")["desc_cargo"].to_dict()

    df["situacao_norm"] = df["situacao"].apply(norm)
    df["sithoje_norm"]  = df["sithoje"].apply(norm)
    situ_up = df["situacao"].astype(str).str.upper()

    df["flag_inss"] = situ_up.str.contains("INSS", na=False, regex=False)

    df["flag_ferias"] = (
        situ_up.str.contains("FERIAS", na=False, regex=False) |
        situ_up.str.contains("FÉRIAS", na=False, regex=False) |
        situ_up.str.contains("F╔RIAS", na=False, regex=False)
    )

    df["flag_suspensao"] = (
        situ_up.str.contains("SUSPENSAO", na=False, regex=False) |
        situ_up.str.contains("SUSPENSÃO", na=False, regex=False) |
        situ_up.str.contains("SUSPENS├O", na=False, regex=False)
    )

    df["flag_falta_injustificada"] = (
        (df["situacao_norm"] == "FALTA") &
        (df["sithoje_norm"] == "TRABALHO")
    )

    df["flag_falta_abonada"] = (
        (df["situacao_norm"] == "FALTA") &
        (df["sithoje_norm"] == "FALTA ABONADA")
    )

    df["flag_falta_justificada"] = (
        (df["situacao_norm"] == "FALTA") &
        (df["sithoje_norm"].isin(["FALTA JUSTIFICADO", "FALTA JUSTIFICADA"]))
    )

    inss_por_re         = df.groupby("re")["flag_inss"].sum().to_dict()
    ferias_por_re       = df.groupby("re")["flag_ferias"].sum().to_dict()
    suspensao_por_re    = df.groupby("re")["flag_suspensao"].sum().to_dict()
    falta_injust_por_re = df.groupby("re")["flag_falta_injustificada"].sum().to_dict()
    falta_abon_por_re   = df.groupby("re")["flag_falta_abonada"].sum().to_dict()
    falta_just_por_re   = df.groupby("re")["flag_falta_justificada"].sum().to_dict()

    linhas_resultado = []

    for re_val, grupo_re in top2_clientes_por_re.groupby("re", sort=False):
        grupo_re = grupo_re.sort_values("qtd", ascending=False)

        top = grupo_re.iloc[0]
        cliente_top = top["cliente"]
        qtd_top = top["qtd"]

        cliente_segundo = ""
        if len(grupo_re) > 1:
            segundo = grupo_re.iloc[1]
            if segundo["qtd"] == qtd_top:
                cliente_segundo = segundo["cliente"]

        linhas_resultado.append({
            "re": re_val,
            "nome": map_re_para_nome.get(re_val, ""),
            "cliente_mais_frequente": cliente_top,
            "cliente_segundo_mais_frequente": cliente_segundo,
            "ESCALA": escala_por_re.get(re_val, ""),
            "CARGO": cargo_por_re.get(re_val, ""),
            "INSS": int(inss_por_re.get(re_val, 0)),
            "Férias": int(ferias_por_re.get(re_val, 0)),
            "Suspensão": int(suspensao_por_re.get(re_val, 0)),
            "Falta Injustificada": int(falta_injust_por_re.get(re_val, 0)),
            "Falta Abonada": int(falta_abon_por_re.get(re_val, 0)),
            "Falta Justificada": int(falta_just_por_re.get(re_val, 0)),
        })

    resultado = pd.DataFrame(linhas_resultado)

    output_path.parent.mkdir(parents=True, exist_ok=True)

    from openpyxl import Workbook
    from openpyxl.styles import Side

    wb = Workbook()
    ws = wb.active
    ws.title = "Resultado"

    headers = list(resultado.columns)
    ws.append(headers)

    for _, row in resultado.iterrows():
        ws.append(list(row.values))

    empty_border = Border()
    for row in ws.iter_rows():
        for cell in row:
            cell.font = Font(name="Calibri", size=8)
            cell.border = empty_border

    header_fill = PatternFill(fill_type="solid", fgColor="8EA9DB")
    for cell in ws[1]:
        cell.font = Font(name="Calibri", size=8, bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = empty_border

    ws.sheet_view.showGridLines = False

    for col_idx, column_cells in enumerate(ws.columns, start=1):
        max_len = 0
        for cell in column_cells:
            value = cell.value
            if value is None:
                continue
            length = len(str(value))
            if length > max_len:
                max_len = length
        adjusted_width = max_len + 4
        ws.column_dimensions[get_column_letter(col_idx)].width = adjusted_width

    for row_idx in range(1, ws.max_row + 1):
        ws.row_dimensions[row_idx].height = 10.2

    wb.save(output_path)

    print()
    print(f'[V2] Arquivo "{output_path.name}" gerado com sucesso!')
    print(f"[V2] Total de linhas no arquivo de entrada: {len(df):,}")
    print(f"[V2] Total de REs distintos no resultado:   {len(resultado):,}")

    return output_path


# =========================================================
# VERSÃO 3 – V1 + MODO STANDALONE PARA TESTES LOCAIS
# =========================================================

def processar_ficha_presenca_v3(input_path, output_path):
    """
    Versão 3:
    - Usa a mesma lógica da V1.
    - Pensada para você testar localmente na mão se quiser,
      mas pode ser usada também pelo sistema web se apontar pra cá.
    """
    return processar_ficha_presenca_v1(input_path, output_path)


# =========================================================
# FUNÇÃO PADRÃO QUE O SISTEMA DO SEU AMIGO VAI CHAMAR
# =========================================================

def processar_ficha_presenca(input_path, output_path):
    """
    Wrapper padrão chamado pelo sistema web.

    Por padrão, usa a VERSÃO 1 (sua lógica original + formatação).
    Se quiser testar a V2 ou V3, basta trocar a linha abaixo.
    """
    return processar_ficha_presenca_v1(input_path, output_path)
    # return processar_ficha_presenca_v2(input_path, output_path)
    # return processar_ficha_presenca_v3(input_path, output_path)


# =========================================================
# MODO DEBUG LOCAL (opcional)
# =========================================================

if __name__ == "__main__":
    fp_teste = Path("FP.xlsx")
    saida_teste = Path("FP_resultado_clientes.xlsx")
    processar_ficha_presenca_v1(fp_teste, saida_teste)
