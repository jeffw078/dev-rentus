from __future__ import annotations

from pathlib import Path
from datetime import datetime
from typing import List, Tuple, Dict, Any
import xml.etree.ElementTree as ET

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font

__all__ = ["process_suprimentos_xml"]  # ajuda import estável

PROCESS_LOGS: List[str] = []


def log(msg: str) -> None:
    print(msg)
    PROCESS_LOGS.append(msg)


def _safe_sheet_name(name: str) -> str:
    bad = r"[]:*?/\\"
    for ch in bad:
        name = name.replace(ch, " ")
    name = " ".join(name.split()).strip()
    return (name or "EMPRESA")[:31]


def parse_xml_bytes(xml_bytes: bytes, source_name: str = "") -> Dict[str, Any]:
    root = ET.fromstring(xml_bytes)

    def get_text(tag: str) -> str:
        el = root.find(tag)
        return (el.text or "").strip() if el is not None else ""

    empresa_id = get_text("id")
    empresa_nome = get_text("nome")
    cnpj = get_text("cnpj")
    data_geracao = get_text("dataGeracao") or datetime.now().strftime("%Y-%m-%d")

    produtos = []
    produtos_el = root.find("produtos")
    if produtos_el is not None:
        for p in produtos_el.findall("produto"):
            def ptext(t: str) -> str:
                el = p.find(t)
                return (el.text or "").strip() if el is not None else ""

            codigo = ptext("codigo")
            descricao = ptext("descricao")
            unidade = ptext("unidade")

            try:
                quantidade = float((ptext("quantidade") or "0").replace(",", "."))
            except Exception:
                quantidade = 0.0

            try:
                valor_unitario = float((ptext("valorUnitario") or "0").replace(",", "."))
            except Exception:
                valor_unitario = 0.0

            total_linha = quantidade * valor_unitario

            produtos.append({
                "empresa_id": empresa_id,
                "empresa_nome": empresa_nome,
                "cnpj": cnpj,
                "data_geracao": data_geracao,
                "origem_arquivo": source_name,
                "produto_codigo": codigo,
                "produto_descricao": descricao,
                "unidade": unidade,
                "quantidade": quantidade,
                "valor_unitario": valor_unitario,
                "total_linha": total_linha,
            })

    return {
        "empresa_id": empresa_id,
        "empresa_nome": empresa_nome,
        "cnpj": cnpj,
        "data_geracao": data_geracao,
        "produtos": produtos,
    }


def process_suprimentos_xml(
    xml_files: List[Tuple[str, bytes]],
    output_dir: Path,
):
    PROCESS_LOGS.clear()
    output_dir.mkdir(parents=True, exist_ok=True)

    if not xml_files:
        raise ValueError("Nenhum XML recebido.")
    if len(xml_files) > 10:
        raise ValueError("Máximo de 10 XML por processamento.")

    log("[INFO] Iniciando Módulo 2 (Suprimentos XML)")
    rows_by_company: Dict[str, List[Dict[str, Any]]] = {}

    total_produtos = 0
    for fname, content in xml_files:
        log(f"[INFO] Lendo XML: {fname}")
        parsed = parse_xml_bytes(content, source_name=fname)
        empresa = parsed.get("empresa_nome") or "EMPRESA"
        prods = parsed.get("produtos", [])
        total_produtos += len(prods)
        rows_by_company.setdefault(empresa, []).extend(prods)

    if total_produtos == 0:
        raise ValueError("Nenhum produto encontrado nos XMLs (verifique <produtos><produto>...)")

    ts = datetime.now().strftime("%Y%m%d-%H%M%S")
    out_path = output_dir / f"modulo2-suprimentos-{ts}.xlsx"

    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        for empresa, rows in rows_by_company.items():
            df = pd.DataFrame(rows)
            # evita erro se vier vazio por alguma razão
            if df.empty:
                continue
            df = df.sort_values(["data_geracao", "produto_descricao"], kind="stable")
            df.to_excel(writer, sheet_name=_safe_sheet_name(empresa), index=False)

    wb = load_workbook(out_path)
    fill = PatternFill("solid", fgColor="333333")
    font = Font(color="FFFFFF", bold=True)

    for ws in wb.worksheets:
        for c in range(1, ws.max_column + 1):
            ws.cell(row=1, column=c).fill = fill
            ws.cell(row=1, column=c).font = font
        ws.freeze_panes = "A2"

    wb.save(out_path)

    log(f"[INFO] XLSX gerado: {out_path.name}")
    return out_path, PROCESS_LOGS
