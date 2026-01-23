# projects/modulo2/service.py

from datetime import date, datetime, timedelta
from typing import Tuple, List, Dict
import xml.etree.ElementTree as ET
import re
import unicodedata
import random

from .config import DEV_MODE
from .preview import consultar_com_auto_recuperacao_nsu
from .preview_cache import get_preview_cache
from .db import (
    get_empresas,
    get_ultimo_nsu,
    atualizar_nsu,
    salvar_xmls_e_nsu,
    listar_pendencias_db,
    listar_postos_db,
    criar_pendencia,
    atualizar_pendencia_com_posto,
    consultar_nfes_por_data,
    salvar_posto,
    get_conn
)

try:
    from .sefaz_client import SEFAZClient, SEFAZ_ENDPOINT
except ImportError:
    SEFAZClient = None
    SEFAZ_ENDPOINT = None

from .rate_limiter import get_rate_limiter, wait_before_sefaz_request

# Importar enriquecimento de CEPs
try:
    from .processar_enriquecimento import processar_enriquecimento_xml
    # DESABILITADO TEMPORARIAMENTE - estava travando importação
    ENRIQUECIMENTO_HABILITADO = False
    print("[SERVICE] Enriquecimento automatico DESABILITADO (pode ser executado manualmente)")
except ImportError:
    ENRIQUECIMENTO_HABILITADO = False
    print("[SERVICE] Modulo de enriquecimento nao disponivel")


# ================================
# NORMALIZAÇÕES (do tratamento)
# ================================

STOPWORDS_POSTO = {
    "DIRET", "DIRETOR", "DIRETORIA", "DIR",
    "ADMIN", "ADMINISTRACAO", "ADM", "UNIDADE"
}


def normalizar_leve(txt) -> str:
    if txt is None:
        return ""
    txt = str(txt)
    txt = unicodedata.normalize("NFKD", txt)
    txt = "".join(c for c in txt if not unicodedata.combining(c))
    return txt.upper().strip()


def normalizar_forte(txt) -> str:
    if txt is None:
        return ""
    txt = str(txt)
    txt = unicodedata.normalize("NFKD", txt)
    txt = "".join(c for c in txt if not unicodedata.combining(c))
    txt = re.sub(r"[^A-Z0-9]", "", txt.upper())
    return txt


def limpar_posto(txt: str) -> str:
    txt = normalizar_leve(txt)
    palavras = txt.split()
    palavras = [p for p in palavras if p not in STOPWORDS_POSTO]
    return " ".join(palavras)


# ================================
# CONSULTA SEFAZ – QUANTIDADE / REAL
# ================================

def consultar_sefaz_quantidade(
    data_ini: date,
    data_fim: date
) -> dict:
    """
    Consulta a SEFAZ para obter quantidade de XMLs disponíveis.
    Em DEV_MODE, retorna dados mockados. Em produção, consulta SEFAZ real.
    """
    
    print(f"[SERVICE] CONSULTA SEFAZ - Data: {data_ini} a {data_fim}")
    
    # ========================================
    # IMPORTAÇÃO DESABILITADA - USE JSON
    # ========================================
    # Sistema configurado para usar APENAS arquivo JSON (produtos_com_posto.json)
    return {
        "status": "error",
        "mensagem": "❌ CONSULTA SEFAZ DESABILITADA",
        "instrucoes": "Use arquivo: produtos_com_posto.json",
        "comando": "python projects/modulo2/importar_json_produtos.py"
    }
    
    # ========================================
    # MODO PRODUÇÃO: Consultar SEFAZ real
    # ========================================
    print(f"[SERVICE] Modo: PRODUÇÃO (SEFAZ REAL)")
    
    try:
        from .db import get_empresas, get_ultimo_nsu
        from .rate_limiter import wait_before_sefaz_request, get_rate_limiter
        
        # Buscar empresas
        empresas = get_empresas()
        if not empresas:
            return {
                "status": "error",
                "total_encontrado": 0,
                "mensagem": "Nenhuma empresa configurada. Verifique certificados/empresas.json"
            }
        
        # Contar XMLs disponíveis fazendo consulta inicial ao SEFAZ
        total_encontrado = 0
        erros = []
        
        # Importar SEFAZClient apenas se necessário (evitar import se não disponível)
        try:
            from .sefaz_client import SEFAZClient
        except ImportError:
            return {
                "status": "error",
                "total_encontrado": 0,
                "mensagem": "SEFAZClient não disponível. Verifique se todas as dependências estão instaladas."
            }
        
        for empresa in empresas:
            cnpj = empresa["cnpj"]
            try:
                # Buscar último NSU
                ultimo_nsu = get_ultimo_nsu(cnpj)
                
                # Rate limiting
                wait_time = wait_before_sefaz_request(cnpj)
                if wait_time > 0:
                    print(f"[SERVICE] Rate limiting: aguardando {wait_time:.1f}s para empresa {cnpj}")
                
                # Preparar certificado
                cert_pfx = empresa.get("cert_pfx") or empresa.get("caminho_certificado")
                cert_senha = empresa.get("cert_senha") or empresa.get("senha_certificado")
                uf = empresa.get("uf", 35)
                endpoint = empresa.get("sefaz_endpoint")
                
                if not cert_pfx or not cert_senha:
                    print(f"[SERVICE] AVISO: Empresa {cnpj} sem certificado configurado")
                    continue
                
                # Criar cliente SEFAZ e consultar
                client = SEFAZClient(
                    cnpj=cnpj,
                    cert_pfx=cert_pfx,
                    cert_senha=cert_senha,
                    endpoint=endpoint,
                    uf=uf
                )
                
                # Fazer consulta para buscar TODOS os XMLs disponíveis com auto-recuperação
                # IMPORTANTE: atualizar_banco=True porque é IMPORTAÇÃO REAL (persiste NSU)
                xmls, maior_nsu = consultar_com_auto_recuperacao_nsu(client, cnpj, ultimo_nsu, max_iteracoes=20, atualizar_banco=True)
                
                # Registrar requisição no rate limiter
                get_rate_limiter().record_request(cnpj)
                
                # Filtrar XMLs por data (se necessário)
                if xmls:
                    xmls_filtrados = []
                    for x in xmls:
                        try:
                            # Extrair data do XML
                            import xml.etree.ElementTree as ET
                            root = ET.fromstring(x["xml"])
                            data_emissao_str = None
                            
                            for elem in root.iter():
                                if elem.tag.endswith("dhEmi"):
                                    data_emissao_str = elem.text
                                    break
                            
                            if data_emissao_str:
                                # Converter data (formato: 2024-01-15T10:30:00-03:00)
                                if "T" in data_emissao_str:
                                    data_emissao = date.fromisoformat(data_emissao_str.split("T")[0])
                                else:
                                    data_emissao = date.fromisoformat(data_emissao_str[:10])
                                
                                # Verificar se está no intervalo
                                if data_ini <= data_emissao <= data_fim:
                                    xmls_filtrados.append(x)
                        except Exception:
                            # Se não conseguir extrair data, incluir (melhor incluir que excluir)
                            xmls_filtrados.append(x)
                    
                    total_encontrado += len(xmls_filtrados)
                    print(f"[SERVICE] Empresa {cnpj}: {len(xmls_filtrados)} XMLs no período")
                else:
                    print(f"[SERVICE] Empresa {cnpj}: Nenhum XML novo encontrado")
                
            except Exception as e:
                erro_msg = f"Erro ao consultar empresa {cnpj}: {str(e)}"
                print(f"[SERVICE] {erro_msg}")
                erros.append(erro_msg)
                continue
        
        if erros and total_encontrado == 0:
            return {
                "status": "error",
                "total_encontrado": 0,
                "mensagem": f"Erro ao consultar SEFAZ: {erros[0]}"
            }
        
        mensagem = f"{total_encontrado} arquivo(s) encontrado(s) no período"
        if erros:
            mensagem += f" ({len(erros)} aviso(s))"
        
        return {
            "status": "ok",
            "total_encontrado": total_encontrado,
            "mensagem": mensagem,
            "avisos": erros if erros else None
        }
        
    except Exception as e:
        print(f"[SERVICE] ERRO CRÍTICO na consulta SEFAZ: {e}")
        import traceback
        traceback.print_exc()
        return {
            "status": "error",
            "total_encontrado": 0,
            "mensagem": f"Erro ao consultar SEFAZ: {str(e)}"
        }


# ================================
# IMPORTAR XMLs DO SEFAZ
# ================================

def _gerar_xml_mock(cnpj: str, nsu: int, data_emissao: date = None) -> str:
    """
    Gera um XML mockado para desenvolvimento.
    Usado quando MODULO2_DEV_MODE=true para simular dados do SEFAZ.
    """
    if data_emissao is None:
        data_emissao = date.today() - timedelta(days=random.randint(1, 30))
    
    # Chave NFe tem 44 dígitos: UF(2) + AAMM(4) + CNPJ(14) + Mod(2) + Série(3) + NNF(9) + tpEmis(1) + cNF(8) + DV(1)
    chave = f"35{data_emissao.strftime('%y%m')}{cnpj}55001{nsu:09d}1{random.randint(10000000, 99999999):08d}{random.randint(0,9)}"
    
    # Nomes mockados de fornecedores
    fornecedores = [
        "FORNECEDOR ABC LTDA",
        "SUPPLY COMPANY XYZ",
        "DISTRIBUIDORA 123",
        "COMERCIAL DEF LTDA"
    ]
    
    # Buscar postos reais do banco
    try:
        postos_db = listar_postos_db()
        if postos_db and len(postos_db) > 0:
            # Selecionar posto aleatório do banco
            posto_selecionado = random.choice(postos_db)
            # IMPORTANTE: usar apenas nomepos no infCpl para identificação funcionar
            # A função identificar_posto busca por nomepos normalizado no índice
            posto_nome = posto_selecionado.get('nomepos', 'POSTO')
            nomecli_posto = posto_selecionado.get('nomecli', 'CLIENTE')
            endereco_posto = posto_selecionado.get('end', 'RUA EXEMPLO')
            cidade_posto = posto_selecionado.get('nomecid', 'SAO PAULO')
            uf_posto = posto_selecionado.get('estado', 'SP') or 'SP'
            cep_posto = posto_selecionado.get('cep', '02000000') or '02000000'
        else:
            # Fallback se não houver postos no banco
            posto_nome = "POSTO CENTRAL"
            nomecli_posto = "CLIENTE"
            endereco_posto = "RUA EXEMPLO"
            cidade_posto = "SAO PAULO"
            uf_posto = "SP"
            cep_posto = "02000000"
    except Exception as e:
        print(f"[MOCK] ERRO ao buscar postos do banco: {e}, usando fallback")
        posto_nome = "POSTO CENTRAL"
        nomecli_posto = "CLIENTE"
        endereco_posto = "RUA EXEMPLO"
        cidade_posto = "SAO PAULO"
        uf_posto = "SP"
        cep_posto = "02000000"
    
    fornecedor = random.choice(fornecedores)
    posto = posto_nome  # Apenas nomepos para ser identificado corretamente
    
    # Gerar valor aleatório
    valor = round(random.uniform(500.0, 50000.0), 2)
    
    # XML mockado básico
    xml = f"""<?xml version="1.0" encoding="UTF-8"?>
<NFe xmlns="http://www.portalfiscal.inf.br/nfe">
  <infNFe Id="NFe{chave}" versao="4.00">
    <ide>
      <cUF>35</cUF>
      <cNF>{random.randint(100000, 999999)}</cNF>
      <mod>55</mod>
      <serie>1</serie>
      <nNF>{nsu}</nNF>
      <dhEmi>{data_emissao.isoformat()}T10:00:00-03:00</dhEmi>
      <dhSaiEnt>{data_emissao.isoformat()}T14:00:00-03:00</dhSaiEnt>
      <tpNF>1</tpNF>
      <idDest>1</idDest>
      <cMunFG>3550308</cMunFG>
      <tpImp>1</tpImp>
      <tpEmis>1</tpEmis>
      <cDV>5</cDV>
      <tpAmb>1</tpAmb>
    </ide>
    <emit>
      <CNPJ>{cnpj}</CNPJ>
      <xNome>{fornecedor}</xNome>
      <xFant>{fornecedor}</xFant>
      <enderEmit>
        <xLgr>RUA EXEMPLO</xLgr>
        <nro>123</nro>
        <xBairro>CENTRO</xBairro>
        <cMun>3550308</cMun>
        <xMun>SAO PAULO</xMun>
        <UF>SP</UF>
        <CEP>01000000</CEP>
      </enderEmit>
    </emit>
    <dest>
      <CNPJ>12345678000190</CNPJ>
      <xNome>{nomecli_posto}</xNome>
      <enderDest>
        <xLgr>{endereco_posto}</xLgr>
        <nro>456</nro>
        <xBairro>BAIRRO</xBairro>
        <cMun>3550308</cMun>
        <xMun>{cidade_posto}</xMun>
        <UF>{uf_posto}</UF>
        <CEP>{cep_posto}</CEP>
      </enderDest>
    </dest>
    <det nItem="1">
      <prod>
        <cProd>PROD001</cProd>
        <cEAN></cEAN>
        <xProd>PRODUTO EXEMPLO</xProd>
        <NCM>99999999</NCM>
        <CFOP>5102</CFOP>
        <uCom>UN</uCom>
        <qCom>10.00</qCom>
        <vUnCom>{valor/10:.2f}</vUnCom>
        <vProd>{valor:.2f}</vProd>
        <cEANTrib></cEANTrib>
        <uTrib>UN</uTrib>
        <qTrib>10.00</qTrib>
        <vUnTrib>{valor/10:.2f}</vUnTrib>
      </prod>
      <imposto>
        <vTotTrib>0.00</vTotTrib>
        <ICMS>
          <ICMS00>
            <orig>0</orig>
            <CST>000</CST>
            <modBC>0</modBC>
            <vBC>{valor:.2f}</vBC>
            <pICMS>18.00</pICMS>
            <vICMS>0.00</vICMS>
          </ICMS00>
        </ICMS>
      </imposto>
    </det>
    <total>
      <ICMSTot>
        <vBC>0.00</vBC>
        <vICMS>0.00</vICMS>
        <vICMSDeson>0.00</vICMSDeson>
        <vFCP>0.00</vFCP>
        <vBCST>0.00</vBCST>
        <vST>0.00</vST>
        <vFCPST>0.00</vFCPST>
        <vFCPSTRet>0.00</vFCPSTRet>
        <vProd>{valor:.2f}</vProd>
        <vFrete>0.00</vFrete>
        <vSeg>0.00</vSeg>
        <vDesc>0.00</vDesc>
        <vII>0.00</vII>
        <vIPI>0.00</vIPI>
        <vIPIDevol>0.00</vIPIDevol>
        <vPIS>0.00</vPIS>
        <vCOFINS>0.00</vCOFINS>
        <vOutro>0.00</vOutro>
        <vNF>{valor:.2f}</vNF>
        <vTotTrib>0.00</vTotTrib>
      </ICMSTot>
    </total>
    <infAdic>
      <infCpl>LOCAL DE ENTREGA: {posto}</infCpl>
    </infAdic>
  </infNFe>
</NFe>"""
    
    return xml


def importar_xmls_sefaz(
    data_ini: date,
    data_fim: date
) -> dict:
    """
    Importa XMLs do SEFAZ para o banco de dados.
    Em DEV_MODE, gera XMLs mockados. Em produção, consulta SEFAZ real.
    Usa NSU incremental para evitar duplicatas.
    """
    
    print(f"[SERVICE] IMPORTAR XMLs SEFAZ - Data: {data_ini} a {data_fim}")
    
    # ========================================
    # MODO DEV: Gerar XMLs mockados
    # ========================================
    # IMPORTAÇÃO DESABILITADA - USE JSON
    # ========================================
    # Sistema configurado para usar APENAS arquivo JSON (produtos_com_posto.json)
    return {
        "status": "error",
        "mensagem": "❌ IMPORTAÇÃO VIA SEFAZ DESABILITADA",
        "instrucoes": "Use arquivo: produtos_com_posto.json",
        "comando": "python projects/modulo2/importar_json_produtos.py"
    }
    
    # ========================================
    # Verificar se existe cache válido do preview
    cache = get_preview_cache()
    cache_data = cache.get()
    
    if cache_data:
        xmls_por_empresa_cache, nsu_por_empresa_cache = cache_data
        print(f"[SERVICE] Cache encontrado! Usando {sum(len(x) for x in xmls_por_empresa_cache.values())} XMLs do preview")
        print(f"[SERVICE] EVITANDO re-consultar SEFAZ (economia de requisicoes)")
    else:
        print(f"[SERVICE] Cache não encontrado ou expirado. Consultando SEFAZ...")
        xmls_por_empresa_cache = None
        nsu_por_empresa_cache = None
    
    try:
        # Buscar empresas do banco
        empresas = get_empresas()
        
        # Se não houver empresas, tentar recarregar do JSON
        if not empresas:
            print("[SERVICE] Nenhuma empresa encontrada no banco. Tentando recarregar do JSON...")
            from .db import seed_empresas_from_json
            # Forçar recarregamento
            seed_empresas_from_json(force=True)
            empresas = get_empresas()
        
        if not empresas:
            return {
                "success": False,
                "error": "Nenhuma empresa configurada. Verifique se certificados/empresas.json existe e tem empresas válidas. Reinicie o servidor se necessário."
            }
        
        print(f"[SERVICE] {len(empresas)} empresa(s) encontrada(s) no banco")
        print(f"[SERVICE] {'='*60}")
        
        total_importado = 0
        total_encontrado = 0
        erros = []
        resumo_empresas = []
        
        for idx, empresa in enumerate(empresas, 1):
            cnpj = empresa["cnpj"]
            print(f"\n[SERVICE] [{idx}/{len(empresas)}] Processando empresa {cnpj}")
            
            try:
                # Buscar último NSU
                ultimo_nsu = get_ultimo_nsu(cnpj)
                print(f"[SERVICE]   - Último NSU conhecido: {ultimo_nsu}")
                
                # CONSULTAR SEFAZ REAL (sem dados mockados)
                cert_pfx = empresa.get("cert_pfx") or empresa.get("caminho_certificado")
                cert_senha = empresa.get("cert_senha") or empresa.get("senha_certificado")
                uf = empresa.get("uf", 43)
                endpoint = empresa.get("sefaz_endpoint") or (SEFAZ_ENDPOINT if SEFAZ_ENDPOINT else None)
                
                if not cert_pfx or not cert_senha:
                    print(f"[SERVICE] AVISO: Empresa {cnpj} sem certificado configurado")
                    erros.append(f"Empresa {cnpj} sem certificado configurado")
                    continue
                
                if not SEFAZClient:
                    print(f"[SERVICE] ERRO: SEFAZClient não disponível")
                    erros.append(f"SEFAZClient não disponível para empresa {cnpj}")
                    continue
                
                try:
                    # Verificar se existe cache para esta empresa
                    if xmls_por_empresa_cache and cnpj in xmls_por_empresa_cache:
                        # Usar XMLs do cache
                        xmls = xmls_por_empresa_cache[cnpj]
                        maior_nsu = nsu_por_empresa_cache[cnpj]
                        print(f"[SERVICE]   - Usando XMLs do CACHE (evitando requisição ao SEFAZ)")
                        print(f"[SERVICE]   - XMLs no cache: {len(xmls)} (NSU até {maior_nsu})")
                    else:
                        # Cache não disponível, consultar SEFAZ
                        # Rate limiting: aguardar antes de consultar SEFAZ
                        wait_time = wait_before_sefaz_request(cnpj)
                        if wait_time > 0:
                            print(f"[SERVICE] Rate limiting aplicado: aguardado {wait_time:.1f}s")
                        
                        # Criar cliente SEFAZ
                        client = SEFAZClient(
                            cnpj=cnpj,
                            cert_pfx=cert_pfx,
                            cert_senha=cert_senha,
                            endpoint=endpoint,
                            uf=uf
                        )
                        
                        # Consultar TODOS os XMLs novos via NSU incremental com auto-recuperação
                        # IMPORTANTE: atualizar_banco=True porque é IMPORTAÇÃO REAL (persiste NSU)
                        print(f"[SERVICE]   - Consultando SEFAZ para TODOS os XMLs novos...")
                        xmls, maior_nsu = consultar_com_auto_recuperacao_nsu(client, cnpj, ultimo_nsu, max_iteracoes=20, atualizar_banco=True)
                        
                        # Registrar requisição no rate limiter
                        get_rate_limiter().record_request(cnpj)
                    
                    xmls_encontrados = len(xmls)
                    total_encontrado += xmls_encontrados
                    
                    print(f"[SERVICE]   - [OK] XMLs ENCONTRADOS no SEFAZ: {xmls_encontrados} (NSU até {maior_nsu})")
                    
                    if xmls:
                        # Validar XMLs antes de salvar (prevenir dados mock)
                        xmls_validos = []
                        xmls_rejeitados = 0
                        try:
                            from .validacao import validar_xml_recebido
                            for x in xmls:
                                is_valid, msg = validar_xml_recebido(x["xml"], x["nsu"])
                                if is_valid:
                                    xmls_validos.append(x)
                                else:
                                    print(f"[SERVICE]   - [VALIDACAO] XML NSU {x['nsu']} rejeitado: {msg}")
                                    xmls_rejeitados += 1
                        except ImportError:
                            # Se módulo de validação não estiver disponível, usar todos
                            xmls_validos = xmls
                        
                        if xmls_rejeitados > 0:
                            print(f"[SERVICE]   - [VALIDACAO] {xmls_rejeitados} XMLs rejeitados por validação (dados mock ou inválidos)")
                        
                        if not xmls_validos:
                            print(f"[SERVICE]   - [VALIDACAO] Nenhum XML válido após validação. Pulando salvamento.")
                            continue
                        
                        # Converter formato para salvar
                        xmls_tuples = [(str(x["nsu"]), x["xml"]) for x in xmls_validos]
                        
                        # Salvar no banco
                        print(f"[SERVICE]   - Salvando {len(xmls_validos)} XMLs válidos no banco de dados...")
                        salvar_xmls_e_nsu(
                            cnpj=cnpj,
                            xmls=xmls_tuples,
                            ultimo_nsu=maior_nsu
                        )
                        
                        # Atualizar contadores (usar apenas XMLs válidos)
                        xmls = xmls_validos
                        
                        # Processar XMLs importados (tratamento)
                        print(f"[SERVICE]   - Processando XMLs (identificação e pendencias)...")
                        for x in xmls:
                            try:
                                processar_xml_e_criar_pendencias(x["xml"])
                            except Exception as e:
                                print(f"[SERVICE]   - [AVISO] ERRO ao processar XML NSU {x['nsu']}: {e}")
                                erros.append(f"Erro ao processar XML NSU {x['nsu']}: {str(e)}")
                        
                        xmls_importados = len(xmls)
                        total_importado += xmls_importados
                        resumo_empresas.append({
                            "cnpj": cnpj,
                            "encontrados": xmls_encontrados,
                            "importados": xmls_importados
                        })
                        print(f"[SERVICE]   - [OK] Empresa {cnpj}: {xmls_importados} XMLs importados com sucesso")
                    else:
                        print(f"[SERVICE]   - [INFO] Nenhum XML novo encontrado para empresa {cnpj}")
                        resumo_empresas.append({
                            "cnpj": cnpj,
                            "encontrados": 0,
                            "importados": 0
                        })
                
                except Exception as e:
                    print(f"[SERVICE] ERRO ao processar empresa {cnpj}: {e}")
                    import traceback
                    traceback.print_exc()
                    erros.append(f"Erro ao processar empresa {cnpj}: {str(e)}")
                    continue
                        
            except Exception as e:
                print(f"[SERVICE] ERRO geral ao processar empresa {cnpj}: {e}")
                import traceback
                traceback.print_exc()
                erros.append(f"Erro geral ao processar empresa {cnpj}: {str(e)}")
                continue
        
        # Resumo final
        print(f"\n[SERVICE] {'='*60}")
        print(f"[SERVICE] RESUMO DA IMPORTAÇÃO:")
        print(f"[SERVICE]   - Total de XMLs ENCONTRADOS no SEFAZ: {total_encontrado}")
        print(f"[SERVICE]   - Total de XMLs IMPORTADOS: {total_importado}")
        if resumo_empresas:
            print(f"[SERVICE]   - Detalhamento por empresa:")
            for resumo in resumo_empresas:
                print(f"[SERVICE]     * {resumo['cnpj']}: {resumo['encontrados']} encontrados -> {resumo['importados']} importados")
        if erros:
            print(f"[SERVICE]   - [AVISO] {len(erros)} aviso(s)/erro(s) durante o processamento")
        print(f"[SERVICE] {'='*60}\n")
        
        if erros and total_importado == 0:
            return {
                "success": False,
                "error": f"Erros durante importação: {'; '.join(erros[:3])}",
                "total": 0
            }
        
        mensagem = f"{total_importado} XMLs importados com sucesso (de {total_encontrado} encontrados no SEFAZ)"
        if erros:
            mensagem += f" ({len(erros)} aviso(s))"
        
        # Limpar cache após importação bem-sucedida
        if total_importado > 0:
            cache.clear()
            print(f"[SERVICE] Cache limpo após importação bem-sucedida")
        
        return {
            "success": True,
            "total": total_importado,
            "total_encontrado": total_encontrado,
            "mensagem": mensagem,
            "erros": erros if erros else None
        }
    
    except Exception as e:
        print(f"[SERVICE] ERRO CRÍTICO na importação: {e}")
        import traceback
        traceback.print_exc()
        return {
            "success": False,
            "error": f"Erro crítico: {str(e)}",
            "total": 0
        }


def importar_xmls_inicial() -> dict:
    """
    Importação inicial (Dia 0): Importa todos os XMLs desde início do ano até hoje.
    Usa NSU incremental em lotes para evitar sobrecarga.
    Processa em lotes de 30 dias com delays entre lotes.
    """
    from .utils import obter_periodo_ano_atual
    
    data_ini, data_fim = obter_periodo_ano_atual()
    
    print(f"[SERVICE] IMPORTAÇÃO INICIAL - Período: {data_ini} até {data_fim}")
    print(f"[SERVICE] Este processo pode demorar várias horas. Processando em lotes...")
    
    # Registrar início da importação no log
    from .db import get_conn, _row_to_dict
    import time
    
    conn = None
    log_id = None
    tempo_inicio = time.time()
    
    try:
        conn = get_conn()
        cur = conn.cursor()
        cur.execute("""
            INSERT INTO modulo2_importacoes_log (
                tipo, data_inicio, data_fim, status, iniciado_em
            )
            VALUES (?, ?, ?, 'em_andamento', datetime('now'))
        """, ("inicial", str(data_ini), str(data_fim)))
        conn.commit()
        log_id = cur.lastrowid
        cur.close()
    except Exception as e:
        print(f"[SERVICE] AVISO: Não foi possível registrar log de importação: {e}")
    
    try:
        # Importar usando a função normal (ela usa NSU incremental automaticamente)
        result = importar_xmls_sefaz(data_ini, data_fim)
        
        # Atualizar log
        if conn and log_id:
            tempo_total = int(time.time() - tempo_inicio)
            cur = conn.cursor()
            cur.execute("""
                UPDATE modulo2_importacoes_log
                SET 
                    total_xmls = ?,
                    xmls_processados = ?,
                    status = ?,
                    mensagem = ?,
                    tempo_execucao_segundos = ?,
                    concluido_em = datetime('now')
                WHERE id = ?
            """, (
                result.get("total", 0),
                result.get("total", 0),
                "concluido" if result.get("success") else "erro",
                result.get("mensagem") or result.get("error", ""),
                tempo_total,
                log_id
            ))
            conn.commit()
            cur.close()
        
        return {
            "success": result.get("success", False),
            "total": result.get("total", 0),
            "mensagem": f"Importação inicial concluída. {result.get('mensagem', '')}",
            "tempo_segundos": int(time.time() - tempo_inicio),
            "periodo": {
                "data_inicio": str(data_ini),
                "data_fim": str(data_fim)
            },
            "erros": result.get("erros")
        }
        
    except Exception as e:
        print(f"[SERVICE] ERRO na importação inicial: {e}")
        import traceback
        traceback.print_exc()
        
        # Atualizar log com erro
        if conn and log_id:
            tempo_total = int(time.time() - tempo_inicio)
            cur = conn.cursor()
            cur.execute("""
                UPDATE modulo2_importacoes_log
                SET 
                    status = 'erro',
                    mensagem = ?,
                    tempo_execucao_segundos = ?,
                    concluido_em = datetime('now')
                WHERE id = ?
            """, (str(e), tempo_total, log_id))
            conn.commit()
            cur.close()
        
        return {
            "success": False,
            "error": f"Erro na importação inicial: {str(e)}",
            "total": 0
        }
    finally:
        if conn:
            conn.close()


def importar_xmls_diario_automatico() -> dict:
    """
    Importação automática diária (executada às 00:00).
    Importa apenas XMLs do dia anterior usando NSU incremental.
    """
    from .utils import obter_periodo_dia_anterior
    
    data_ini, data_fim = obter_periodo_dia_anterior()
    
    print(f"\n{'='*60}")
    print(f"[SERVICE] IMPORTAÇÃO DIÁRIA AUTOMÁTICA INICIADA")
    print(f"[SERVICE] Data de referência: {data_ini}")
    print(f"[SERVICE] Modo: Busca incremental por NSU (apenas XMLs novos)")
    print(f"{'='*60}")
    
    # Registrar início da importação no log
    from .db import get_conn
    import time
    
    conn = None
    log_id = None
    tempo_inicio = time.time()
    
    try:
        conn = get_conn()
        cur = conn.cursor()
        cur.execute("""
            INSERT INTO modulo2_importacoes_log (
                tipo, data_inicio, data_fim, status, iniciado_em
            )
            VALUES (?, ?, ?, 'em_andamento', datetime('now'))
        """, ("diaria", str(data_ini), str(data_fim)))
        conn.commit()
        log_id = cur.lastrowid
        cur.close()
    except Exception as e:
        print(f"[SERVICE] AVISO: Não foi possível registrar log: {e}")
    
    try:
        # Importar usando NSU incremental (apenas XMLs novos)
        result = importar_xmls_sefaz(data_ini, data_fim)
        
        # Log adicional após importação
        total_encontrado = result.get("total_encontrado", result.get("total", 0))
        total_importado = result.get("total", 0)
        
        print(f"\n[SERVICE] {'='*60}")
        print(f"[SERVICE] IMPORTAÇÃO DIÁRIA AUTOMÁTICA CONCLUÍDA")
        print(f"[SERVICE]   - XMLs encontrados no SEFAZ: {total_encontrado}")
        print(f"[SERVICE]   - XMLs importados: {total_importado}")
        
        # Contar XMLs identificados e pendentes
        if conn:
            cur = conn.cursor()
            cur.execute("""
                SELECT 
                    COUNT(CASE WHEN posto_id IS NOT NULL THEN 1 END) as identificados,
                    COUNT(CASE WHEN posto_id IS NULL THEN 1 END) as pendentes
                FROM modulo2_nfe
                WHERE data_emissao = ?
            """, (str(data_ini),))
            row = cur.fetchone()
            cur.close()
            
            identificados = row[0] if row else 0
            pendentes = row[1] if row else 0
        else:
            identificados = 0
            pendentes = 0
        
        print(f"[SERVICE]   - XMLs identificados (com posto): {identificados}")
        print(f"[SERVICE]   - XMLs pendentes (sem posto): {pendentes}")
        
        # Atualizar log
        if conn and log_id:
            tempo_total = int(time.time() - tempo_inicio)
            print(f"[SERVICE]   - Tempo de execução: {tempo_total}s")
            print(f"[SERVICE] {'='*60}\n")
            cur = conn.cursor()
            cur.execute("""
                UPDATE modulo2_importacoes_log
                SET 
                    total_xmls = ?,
                    xmls_processados = ?,
                    xmls_identificados = ?,
                    xmls_pendentes = ?,
                    status = ?,
                    mensagem = ?,
                    tempo_execucao_segundos = ?,
                    concluido_em = datetime('now')
                WHERE id = ?
            """, (
                result.get("total", 0),
                result.get("total", 0),
                identificados,
                pendentes,
                "concluido" if result.get("success") else "erro",
                result.get("mensagem") or result.get("error", ""),
                tempo_total,
                log_id
            ))
            conn.commit()
            cur.close()
        
        return {
            "success": result.get("success", False),
            "total": result.get("total", 0),
            "identificados": identificados,
            "pendentes": pendentes,
            "mensagem": result.get("mensagem", ""),
            "tempo_segundos": int(time.time() - tempo_inicio)
        }
        
    except Exception as e:
        print(f"[SERVICE] ERRO na importação diária automática: {e}")
        import traceback
        traceback.print_exc()
        
        # Atualizar log com erro
        if conn and log_id:
            tempo_total = int(time.time() - tempo_inicio)
            cur = conn.cursor()
            cur.execute("""
                UPDATE modulo2_importacoes_log
                SET 
                    status = 'erro',
                    mensagem = ?,
                    tempo_execucao_segundos = ?,
                    concluido_em = datetime('now')
                WHERE id = ?
            """, (str(e), tempo_total, log_id))
            conn.commit()
            cur.close()
        
        return {
            "success": False,
            "error": f"Erro na importação diária: {str(e)}",
            "total": 0
        }
    finally:
        if conn:
            conn.close()


# ================================
# TRATAMENTO DE XMLs
# ================================

def processar_xml_e_criar_pendencias(xml_string: str):
    """
    Processa um XML e tenta identificar o posto de trabalho.
    Se não conseguir, cria uma pendência.
    """
    try:
        root = ET.fromstring(xml_string)
        
        # Extrair informações básicas
        chave = extrair_chave_nfe(root)
        if not chave:
            return
        
        valor_total = extrair_valor_total(root)
        fornecedor = extrair_fornecedor(root)
        infcpl = extrair_infCpl(root)
        enderDest = extrair_enderDest(root)
        
        # Buscar NFe no banco
        conn = None
        try:
            conn = get_conn()
            cur = conn.cursor()
            
            cur.execute("SELECT id FROM modulo2_nfe WHERE chave_acesso = ?", (chave,))
            nfe_row = cur.fetchone()
            
            if not nfe_row:
                cur.close()
                conn.close()
                return  # NFe não encontrada no banco
            
            nfe_id = nfe_row[0]
            cur.close()
            
            # ============================================
            # ENRIQUECIMENTO DE CEPs (se habilitado)
            # ============================================
            if ENRIQUECIMENTO_HABILITADO:
                try:
                    resultado_enriq = processar_enriquecimento_xml(
                        xml_string=xml_string,
                        nfe_id=nfe_id,
                        chave_nfe=chave
                    )
                    
                    if resultado_enriq.get('cep_atualizado'):
                        print(f"[ENRIQUECIMENTO] ✅ {resultado_enriq['mensagem']}")
                    elif resultado_enriq.get('posto_sugerido'):
                        print(f"[ENRIQUECIMENTO] 📝 {resultado_enriq['mensagem']}")
                        
                except Exception as e:
                    # Não deixar falha de enriquecimento quebrar o fluxo principal
                    print(f"[ENRIQUECIMENTO] ⚠️  Erro (não crítico): {e}")
            
            # Tentar identificar posto
            posto = identificar_posto(infcpl, enderDest)
            
            if posto:
                # Atualizar NFe com posto_id
                cur = conn.cursor()
                cur.execute("""
                    UPDATE modulo2_nfe 
                    SET posto_id = ?, status = 'identificado', updated_at = datetime('now')
                    WHERE id = ?
                """, (posto["id"], nfe_id))
                conn.commit()
                cur.close()
                
                print(f"[TRATAMENTO] NFe {chave} identificada com posto {posto.get('nomepos', posto.get('nome'))}")
            else:
                # Criar pendência
                motivo = "Não foi possível identificar posto de trabalho automaticamente"
                criar_pendencia(
                    nfe_id=nfe_id,
                    chave_nfe=chave,
                    valor=valor_total or 0,
                    fornecedor=fornecedor or "DESCONHECIDO",
                    motivo=motivo
                )
                print(f"[TRATAMENTO] Pendência criada para NFe {chave}")
            
        finally:
            if conn:
                conn.close()
        
    except Exception as e:
        print(f"[TRATAMENTO] ERRO ao processar XML: {e}")


def identificar_posto(infcpl: str, enderDest: dict) -> dict:
    """
    Tenta identificar o posto de trabalho usando as regras do tratamento.
    Retorna dict com informações do posto ou None se não encontrar.
    
    VERSÃO MELHORADA: Agora tenta identificar por enderDest mesmo se infCpl vazio!
    """
    # ✅ NÃO retornar None se infCpl vazio - tentar por enderDest também
    
    # Carregar postos do banco
    postos = listar_postos_db()
    if not postos:
        return None
    
    # Criar índices (similar ao código original)
    idx_postos = {}
    idx_cep = {}
    
    for p in postos:
        nome_limpo = limpar_posto(p.get("nomepos", ""))
        idx_postos[normalizar_forte(nome_limpo)] = p
        
        cep = str(p.get("cep", "")).zfill(8) if p.get("cep") else ""
        if cep and cep != "00000000":
            idx_cep.setdefault(cep, []).append(p)
    
    # ============================================
    # FASE 1: TENTAR IDENTIFICAR POR infCpl
    # ============================================
    if infcpl:
        # REGRA 1: LOCAL DE ENTREGA
        # Normalizar infCpl para busca (mas manter original para regex)
        infcpl_normalizado = normalizar_leve(infcpl) if infcpl else ""
        
        if "LOCAL DE ENTREGA" in infcpl_normalizado:
            # Usar regex case-insensitive no texto original
            m = re.search(r"LOCAL\s+DE\s+ENTREGA\s*:\s*([^;]+)", infcpl, re.IGNORECASE)
            if m:
                posto_texto = m.group(1).strip()
                
                # Tentar encontrar exatamente no índice
                posto_limpo = limpar_posto(posto_texto)
                chave_norm = normalizar_forte(posto_limpo)
                posto = idx_postos.get(chave_norm)
                if posto:
                    return posto
                
                # Se não encontrou exato, tentar buscar por parte do texto (caso tenha "nomecli - nomepos")
                # Extrair última parte após "-" ou último token
                if " - " in posto_texto or " -" in posto_texto or "- " in posto_texto:
                    # Tentar pegar a parte após o último "-"
                    partes = re.split(r'\s*-\s*', posto_texto)
                    if len(partes) > 1:
                        # Pegar última parte (geralmente é o nomepos)
                        ultima_parte = partes[-1].strip()
                        posto_limpo = limpar_posto(ultima_parte)
                        chave_norm = normalizar_forte(posto_limpo)
                        posto = idx_postos.get(chave_norm)
                        if posto:
                            return posto
                
                # Tentar buscar parcialmente nos índices (buscar por substring)
                posto_texto_norm = normalizar_forte(posto_limpo)
                melhor_match = None
                melhor_score = 0
                for chave_idx, posto_ref in idx_postos.items():
                    # Verificar se o nome do posto está contido no texto extraído ou vice-versa
                    if chave_idx and posto_texto_norm:
                        # Verificar substring (texto do XML contém nome do posto ou vice-versa)
                        if len(chave_idx) >= 5 and len(posto_texto_norm) >= 5:
                            # Calcular similaridade simples
                            if chave_idx in posto_texto_norm or posto_texto_norm in chave_idx:
                                # Usar o maior match
                                score = min(len(chave_idx), len(posto_texto_norm))
                                if score > melhor_score:
                                    melhor_score = score
                                    melhor_match = posto_ref
                
                if melhor_match:
                    return melhor_match
        
        # REGRA 2: SME
        if infcpl.startswith("SME"):
            m = re.search(r"POSTO\s*\d+\s*-\s*([^;]+)", infcpl)
            if m:
                posto_limpo = limpar_posto(m.group(1))
                chave_norm = normalizar_forte(posto_limpo)
                posto = idx_postos.get(chave_norm)
                if posto:
                    return posto
    
    # ============================================
    # FASE 2: TENTAR IDENTIFICAR POR enderDest
    # ✅ Sempre tentar, mesmo se infCpl vazio!
    # ============================================
    if enderDest:
        cep = enderDest.get("CEP", "")
        if cep:
            cep_limpo = cep.zfill(8)
            
            # ✅ NOVA REGRA: Estratégia 1: CEP exato
            # Se houver MÚLTIPLOS postos no mesmo CEP, tentar desempatar por infCpl/endereço
            if cep_limpo in idx_cep:
                postos_no_cep = idx_cep[cep_limpo]
                
                # Se houver apenas 1 posto, usar ele
                if len(postos_no_cep) == 1:
                    print(f"[IDENTIFICACAO] ✅ Posto único identificado por CEP exato: {cep_limpo}")
                    return postos_no_cep[0]
                
                # ✅ MÚLTIPLOS POSTOS NO MESMO CEP - Tentar desempatar por infCpl
                print(f"[IDENTIFICACAO] ⚠️  {len(postos_no_cep)} postos com mesmo CEP {cep_limpo} - tentando desempatar...")
                
                # Tentar desempatar usando infCpl (buscar nome do posto ou endereço específico)
                if infcpl:
                    melhor_match = None
                    melhor_score = 0
                    
                    for posto in postos_no_cep:
                        score = 0
                        
                        # Match por nome do posto
                        nome_posto_norm = normalizar_forte(posto.get("nomepos", ""))
                        if nome_posto_norm and len(nome_posto_norm) >= 5:
                            if nome_posto_norm in normalizar_forte(infcpl):
                                score += 100  # Match forte por nome do posto
                        
                        # Match por nome do cliente
                        nome_cliente_norm = normalizar_forte(posto.get("nomecli", ""))
                        if nome_cliente_norm and len(nome_cliente_norm) >= 5:
                            if nome_cliente_norm in normalizar_forte(infcpl):
                                score += 50  # Match médio por nome do cliente
                        
                        # Match por endereço no infCpl
                        end_posto_norm = normalizar_forte(posto.get("end", ""))
                        if end_posto_norm and len(end_posto_norm) >= 10:
                            if end_posto_norm in normalizar_forte(infcpl):
                                score += 70  # Match forte por endereço
                        
                        # Match por bairro
                        bairro_posto_norm = normalizar_forte(posto.get("bairro", ""))
                        if bairro_posto_norm and len(bairro_posto_norm) >= 5:
                            if bairro_posto_norm in normalizar_forte(infcpl):
                                score += 30  # Match fraco por bairro
                        
                        if score > melhor_score:
                            melhor_score = score
                            melhor_match = posto
                    
                    # Se conseguiu um match com score significativo, usar
                    if melhor_match and melhor_score >= 50:
                        print(f"[IDENTIFICACAO] ✅ Posto desempatado por infCpl (score: {melhor_score}) - {melhor_match.get('nomepos')}")
                        return melhor_match
                
                # Se não conseguiu desempatar por infCpl, tentar por endereço completo
                logradouro = enderDest.get("xLgr", "").upper()
                numero = enderDest.get("nro", "").upper()
                
                if logradouro:
                    melhor_match = None
                    melhor_score = 0
                    
                    for posto in postos_no_cep:
                        end_posto = (posto.get("end") or "").upper()
                        score = 0
                        
                        # Match por logradouro
                        if logradouro in end_posto or end_posto in logradouro:
                            score += 50
                        
                        # Match por número
                        if numero and numero in end_posto:
                            score += 50
                        
                        if score > melhor_score:
                            melhor_score = score
                            melhor_match = posto
                    
                    if melhor_match and melhor_score >= 50:
                        print(f"[IDENTIFICACAO] ✅ Posto desempatado por endereço (score: {melhor_score}) - {melhor_match.get('nomepos')}")
                        return melhor_match
                
                # ⚠️ Não conseguiu desempatar - NÃO identificar (avançar para próxima estratégia)
                print(f"[IDENTIFICACAO] ⚠️  Não foi possível desempatar {len(postos_no_cep)} postos - avançando para próxima estratégia")
                # NÃO retornar nada aqui - deixar tentar outras estratégias
            
            # ✅ MELHORIA: Estratégia 2: CEP com diferença mínima (±50)
            # Também aplica regra de desempate se houver múltiplos CEPs próximos
            try:
                cep_num = int(cep_limpo)
                candidatos = []  # Lista de (posto, diferenca)
                
                for cep_cadastrado, postos_lista in idx_cep.items():
                    try:
                        cep_cad_num = int(cep_cadastrado)
                        diferenca = abs(cep_num - cep_cad_num)
                        
                        if diferenca <= 50:
                            # Adicionar todos os postos deste CEP como candidatos
                            for posto in postos_lista:
                                candidatos.append((posto, diferenca))
                    except:
                        pass
                
                if candidatos:
                    # Ordenar por diferença (menor primeiro)
                    candidatos.sort(key=lambda x: x[1])
                    menor_diferenca = candidatos[0][1]
                    
                    # Pegar apenas candidatos com a menor diferença
                    melhores_candidatos = [p for p, d in candidatos if d == menor_diferenca]
                    
                    # Se houver apenas 1 candidato com menor diferença, usar
                    if len(melhores_candidatos) == 1:
                        print(f"[IDENTIFICACAO] ✅ Posto identificado por CEP próximo (diferença: {menor_diferenca})")
                        return melhores_candidatos[0]
                    
                    # Se houver múltiplos candidatos, tentar desempatar por infCpl
                    if len(melhores_candidatos) > 1 and infcpl:
                        melhor_match = None
                        melhor_score = 0
                        
                        for posto in melhores_candidatos:
                            score = 0
                            
                            nome_posto_norm = normalizar_forte(posto.get("nomepos", ""))
                            if nome_posto_norm and len(nome_posto_norm) >= 5:
                                if nome_posto_norm in normalizar_forte(infcpl):
                                    score += 100
                            
                            nome_cliente_norm = normalizar_forte(posto.get("nomecli", ""))
                            if nome_cliente_norm and len(nome_cliente_norm) >= 5:
                                if nome_cliente_norm in normalizar_forte(infcpl):
                                    score += 50
                            
                            if score > melhor_score:
                                melhor_score = score
                                melhor_match = posto
                        
                        if melhor_match and melhor_score >= 50:
                            print(f"[IDENTIFICACAO] ✅ Posto desempatado por infCpl em CEP próximo (dif: {menor_diferenca}, score: {melhor_score})")
                            return melhor_match
                    
                    # Se não conseguiu desempatar mas só há candidatos com diferença mínima, usar o primeiro
                    if len(melhores_candidatos) > 1:
                        print(f"[IDENTIFICACAO] ⚠️  {len(melhores_candidatos)} postos com CEP próximo (dif: {menor_diferenca}) - usando primeiro")
                    
                    return melhores_candidatos[0]
            except:
                pass
            
            # Estratégia 3: Prefixo CEP (5 primeiros dígitos)
            # Também aplica regra de desempate
            cep5 = cep_limpo[:5]
            candidatos_prefixo = []
            
            for c, lst in idx_cep.items():
                if c.startswith(cep5):
                    candidatos_prefixo.extend(lst)
            
            if candidatos_prefixo:
                # Se houver apenas 1, usar
                if len(candidatos_prefixo) == 1:
                    print(f"[IDENTIFICACAO] ✅ Posto identificado por prefixo CEP: {cep5}")
                    return candidatos_prefixo[0]
                
                # Se houver múltiplos, tentar desempatar por infCpl
                if len(candidatos_prefixo) > 1 and infcpl:
                    melhor_match = None
                    melhor_score = 0
                    
                    for posto in candidatos_prefixo:
                        score = 0
                        
                        nome_posto_norm = normalizar_forte(posto.get("nomepos", ""))
                        if nome_posto_norm and len(nome_posto_norm) >= 5:
                            if nome_posto_norm in normalizar_forte(infcpl):
                                score += 100
                        
                        nome_cliente_norm = normalizar_forte(posto.get("nomecli", ""))
                        if nome_cliente_norm and len(nome_cliente_norm) >= 5:
                            if nome_cliente_norm in normalizar_forte(infcpl):
                                score += 50
                        
                        if score > melhor_score:
                            melhor_score = score
                            melhor_match = posto
                    
                    if melhor_match and melhor_score >= 50:
                        print(f"[IDENTIFICACAO] ✅ Posto desempatado por infCpl em prefixo CEP (score: {melhor_score})")
                        return melhor_match
                
                # Se não conseguiu desempatar, usar o primeiro
                print(f"[IDENTIFICACAO] ⚠️  {len(candidatos_prefixo)} postos com prefixo CEP {cep5} - usando primeiro")
                return candidatos_prefixo[0]
        
        # ✅ MELHORIA: Estratégia 4: Matching por endereço completo (fuzzy)
        logradouro = enderDest.get("xLgr", "").upper()
        numero = enderDest.get("nro", "").upper()
        bairro = enderDest.get("xBairro", "").upper()
        
        if logradouro:
            melhor_match = None
            melhor_score = 0
            
            for posto in postos:
                end_posto = (posto.get("end") or "").upper()
                bairro_posto = (posto.get("bairro") or "").upper()
                
                score = 0
                
                # Verificar logradouro (peso maior)
                if logradouro in end_posto or end_posto in logradouro:
                    score += 50
                
                # Verificar número (se tiver)
                if numero and numero in end_posto:
                    score += 30
                
                # Verificar bairro
                if bairro and bairro_posto and bairro in bairro_posto:
                    score += 20
                
                if score > melhor_score and score >= 50:  # Mínimo 50 pontos
                    melhor_score = score
                    melhor_match = posto
            
            if melhor_match:
                print(f"[IDENTIFICACAO] ✅ Posto identificado por endereço (score: {melhor_score})")
                return melhor_match
    
    return None


# ================================
# EXTRAÇÃO DE DADOS XML
# ================================

def extrair_chave_nfe(root) -> str:
    """Extrai a chave de acesso da NF-e"""
    for elem in root.iter():
        if elem.tag.endswith("infNFe"):
            chave = elem.attrib.get("Id", "").replace("NFe", "").replace("NFE", "")
            return chave
    return ""


def extrair_valor_total(root) -> float:
    """Extrai o valor total da NF-e"""
    for elem in root.iter():
        if elem.tag.endswith("vNF"):
            try:
                return float(elem.text or 0)
            except:
                pass
    return 0.0


def extrair_fornecedor(root) -> str:
    """Extrai o nome do fornecedor (emitente)"""
    for elem in root.iter():
        if elem.tag.endswith("emit"):
            for child in elem:
                if child.tag.endswith("xNome"):
                    return child.text or "DESCONHECIDO"
    return "DESCONHECIDO"


def extrair_infCpl(root) -> str:
    """Extrai informações complementares (retorna texto original, não normalizado)"""
    for elem in root.iter():
        if elem.tag.endswith("infCpl"):
            texto = elem.text or ""
            # Retornar texto original para permitir regex funcionar corretamente
            return texto.strip()
    return ""


def extrair_enderDest(root) -> dict:
    """Extrai endereço do destinatário"""
    for elem in root.iter():
        if elem.tag.endswith("enderDest"):
            result = {}
            for child in elem:
                if child.tag.endswith("xLgr"):
                    result["xLgr"] = child.text or ""
                elif child.tag.endswith("nro"):
                    result["nro"] = child.text or ""
                elif child.tag.endswith("xBairro"):
                    result["xBairro"] = child.text or ""
                elif child.tag.endswith("xMun"):
                    result["xMun"] = child.text or ""
                elif child.tag.endswith("UF"):
                    result["UF"] = child.text or ""
                elif child.tag.endswith("CEP"):
                    result["CEP"] = re.sub(r"\D", "", child.text or "")
            return result if result else None
    return None


# ================================
# PENDÊNCIAS
# ================================

def listar_pendencias(limit: int = 500, data_ini: date = None, data_fim: date = None) -> List[dict]:
    """Lista pendências de identificação com filtro opcional por data de emissão"""
    print(f"[SERVICE] listar_pendencias - limit: {limit}, data_ini: {data_ini}, data_fim: {data_fim}")
    return listar_pendencias_db(limit, data_ini, data_fim)


def identificar_pendencia(pendencia_id: int, cliente_id: int, posto_id: int) -> dict:
    """
    Identifica uma pendência associando cliente e posto.
    """
    # Buscar informações do posto
    postos = listar_postos_db()
    posto = next((p for p in postos if (p.get("id") == posto_id or str(p.get("id")) == str(posto_id))), None)
    
    if not posto:
        return {
            "success": False,
            "error": "Posto não encontrado"
        }
    
    cliente_nome = posto.get("nomecli", "")
    
    # Atualizar pendência
    sucesso = atualizar_pendencia_com_posto(pendencia_id, posto_id, cliente_nome)
    
    if sucesso:
        return {
            "success": True,
            "mensagem": "NFe identificada com sucesso"
        }
    else:
        return {
            "success": False,
            "error": "Erro ao atualizar pendência"
        }


# ================================
# POSTOS DE TRABALHO
# ================================

def listar_postos() -> List[dict]:
    """Lista todos os postos de trabalho"""
    print("[SERVICE] listar_postos")
    return listar_postos_db()


def listar_gastos_por_posto(data_ini: date = None, data_fim: date = None, cliente_filtro: str = None) -> List[dict]:
    """
    Lista gastos por cliente (agregado por nomecli) com filtro opcional por data.
    Retorna dados agregados para o gráfico com orçado e realizado.
    
    Args:
        data_ini: Data inicial para filtrar NFes (opcional)
        data_fim: Data final para filtrar NFes (opcional)
        cliente_filtro: Nome do cliente para filtrar (opcional, usado ao clicar no gráfico)
    
    Returns:
        Lista de dicts com: nomecli, orcado, realizado, status (orcado - realizado)
    """
    conn = None
    try:
        conn = get_conn()
        cur = conn.cursor()
        from .db import _row_to_dict
        
        # ============================================
        # 1. BUSCAR ORÇADO POR CLIENTE
        # Soma de valor_orcado de todos os postos do cliente
        # ============================================
        orcado_query = """
            SELECT 
                nomecli,
                COALESCE(SUM(valor_orcado), 0) as total_orcado,
                COUNT(*) as total_postos_cliente
            FROM modulo2_postos_trabalho
            WHERE nomecli IS NOT NULL AND nomecli != ''
        """
        
        if cliente_filtro:
            orcado_query += " AND nomecli = ?"
            cur.execute(orcado_query + " GROUP BY nomecli", (cliente_filtro,))
        else:
            cur.execute(orcado_query + " GROUP BY nomecli")
        
        orcado_rows = cur.fetchall()
        orcado_map = {}
        for row in orcado_rows:
            row_dict = _row_to_dict(row)
            nomecli = row_dict.get("nomecli", "")
            if nomecli:
                orcado_map[nomecli] = {
                    "orcado": float(row_dict.get("total_orcado", 0) or 0),
                    "total_postos_cliente": row_dict.get("total_postos_cliente", 0)
                }
        
        # ============================================
        # 2. BUSCAR REALIZADO POR CLIENTE
        # Soma de valor_total das NFes agrupado por cliente
        # ============================================
        data_conditions = []
        params = []
        
        if data_ini:
            data_conditions.append("n.data_emissao >= ?")
            params.append(str(data_ini))
        
        if data_fim:
            data_conditions.append("n.data_emissao <= ?")
            params.append(str(data_fim))
        
        data_where = " AND " + " AND ".join(data_conditions) if data_conditions else ""
        
        cliente_filter_clause = ""
        if cliente_filtro:
            cliente_filter_clause = " AND pt.nomecli = ?"
            params.append(cliente_filtro)
        
        query = f"""
            SELECT 
                COALESCE(pt.nomecli, 'Não identificado') as nomecli,
                COALESCE(SUM(n.valor_total), 0) as total_realizado,
                COUNT(DISTINCT n.id) as total_nfes,
                COUNT(DISTINCT CASE WHEN pt.id IS NOT NULL THEN pt.id END) as total_postos_nfe
            FROM modulo2_nfe n
            LEFT JOIN modulo2_postos_trabalho pt ON n.posto_id = pt.id
            WHERE 1=1 {data_where} {cliente_filter_clause}
            GROUP BY COALESCE(pt.nomecli, 'Não identificado')
            ORDER BY total_realizado DESC
        """
        
        cur.execute(query, params)
        rows = cur.fetchall()
        
        # ============================================
        # 3. COMBINAR ORÇADO + REALIZADO
        # ============================================
        result = []
        clientes_processados = set()
        
        for row in rows:
            row_dict = _row_to_dict(row)
            nomecli = row_dict.get("nomecli", "") or "Não identificado"
            clientes_processados.add(nomecli)
            
            orcado_info = orcado_map.get(nomecli, {"orcado": 0.0, "total_postos_cliente": 0})
            orcado = orcado_info["orcado"]
            realizado = float(row_dict.get("total_realizado", 0) or 0)
            status = orcado - realizado  # Positivo = dentro do orçado, Negativo = acima
            
            result.append({
                "nomecli": nomecli,
                "nome": nomecli,
                "orcado": orcado,
                "realizado": realizado,
                "valor": realizado,
                "status": status,
                "total_nfes": row_dict.get("total_nfes", 0),
                "total_postos": row_dict.get("total_postos_nfe", 0),
                "total_postos_cliente": orcado_info["total_postos_cliente"]
            })
        
        # ============================================
        # 4. INCLUIR CLIENTES COM ORÇADO MAS SEM NFE
        # ============================================
        if not cliente_filtro:
            for nomecli, orcado_info in orcado_map.items():
                if nomecli not in clientes_processados and orcado_info["orcado"] > 0:
                    result.append({
                        "nomecli": nomecli,
                        "nome": nomecli,
                        "orcado": orcado_info["orcado"],
                        "realizado": 0.0,
                        "valor": 0.0,
                        "status": orcado_info["orcado"],
                        "total_nfes": 0,
                        "total_postos": 0,
                        "total_postos_cliente": orcado_info["total_postos_cliente"]
                    })
        
        cur.close()
        return result
        
    except Exception as e:
        print(f"[SERVICE] ERRO ao listar gastos por posto: {e}")
        import traceback
        traceback.print_exc()
        return []
    finally:
        if conn:
            conn.close()


def listar_clientes() -> List[dict]:
    """Lista clientes únicos (agrupados por nomecli)"""
    postos = listar_postos_db()
    clientes = {}
    
    for posto in postos:
        nomecli = posto.get("nomecli", "")
        if nomecli and nomecli not in clientes:
            clientes[nomecli] = {
                "id": nomecli,  # Usando nome como ID por enquanto
                "nome": nomecli
            }
    
    return list(clientes.values())


# ================================
# EXPORTAÇÃO EXCEL
# ================================

def exportar_nfes_excel(
    data_inicio: date = None,
    data_fim: date = None,
    apenas_pendentes: bool = False
) -> str:
    """
    Exporta NFes do banco para arquivo Excel.
    
    Campos incluídos:
    - Chave da NF
    - Número da NF
    - NSU vinculado
    - Data da NF
    - Fornecedor
    - Valor da NF
    - Impostos (ICMS, IPI, PIS, COFINS)
    - Informações de destinatário
    - Informações complementares
    - Status (identificada ou não)
    - Motivo (se não identificada)
    
    Args:
        data_inicio: Data inicial do filtro
        data_fim: Data final do filtro
        apenas_pendentes: Se True, exporta apenas pendentes
    
    Returns:
        Caminho do arquivo Excel gerado
    """
    import json
    from pathlib import Path
    from datetime import datetime as dt
    
    try:
        import pandas as pd
        from openpyxl import load_workbook
        from openpyxl.styles import PatternFill, Font, Alignment
    except ImportError:
        raise RuntimeError("pandas e openpyxl são necessários. Instale com: pip install pandas openpyxl")
    
    # Definir período padrão
    if not data_inicio:
        data_inicio = date(2026, 1, 1)
    if not data_fim:
        data_fim = date.today()
    
    print(f"[EXPORTACAO] Exportando NFes de {data_inicio} a {data_fim}")
    
    # Buscar NFes do banco
    conn = get_conn()
    cur = conn.cursor()
    
    query = """
        SELECT 
            n.chave_acesso,
            n.nsu,
            n.data_emissao,
            n.valor_total,
            n.nome_emitente as fornecedor,
            n.cnpj_emitente,
            n.cnpj_destinatario,
            n.nome_destinatario,
            n.endereco_entrega,
            n.info_adicional,
            n.status,
            n.xml,
            p.motivo as motivo_pendencia,
            pt.nomecli as cliente_identificado,
            pt.nomepos as posto_identificado
        FROM modulo2_nfe n
        LEFT JOIN modulo2_pendencias p ON p.chave_nfe = n.chave_acesso AND p.status = 'pendente'
        LEFT JOIN modulo2_postos_trabalho pt ON pt.id = n.posto_id
        WHERE 1=1
    """
    
    params = []
    
    if data_inicio:
        query += " AND (n.data_emissao >= ? OR n.data_emissao IS NULL)"
        params.append(str(data_inicio))
    
    if data_fim:
        query += " AND (n.data_emissao <= ? OR n.data_emissao IS NULL)"
        params.append(str(data_fim))
    
    if apenas_pendentes:
        query += " AND n.status = 'pendente'"
    
    query += " ORDER BY n.data_emissao, n.nsu"
    
    cur.execute(query, params)
    rows = cur.fetchall()
    cur.close()
    conn.close()
    
    print(f"[EXPORTACAO] {len(rows)} NFes encontradas")
    
    # Função para extrair número da NF e impostos do XML
    def extrair_dados_xml(xml_str):
        numero_nf = ""
        impostos = {"icms": 0, "ipi": 0, "pis": 0, "cofins": 0}
        
        if not xml_str:
            return numero_nf, impostos
        
        try:
            root = ET.fromstring(xml_str)
            
            for elem in root.iter():
                tag = elem.tag.split("}")[-1] if "}" in elem.tag else elem.tag
                
                if tag == "nNF":
                    numero_nf = elem.text or ""
                elif tag == "vICMS":
                    try:
                        impostos["icms"] = float(elem.text or 0)
                    except:
                        pass
                elif tag == "vIPI":
                    try:
                        impostos["ipi"] = float(elem.text or 0)
                    except:
                        pass
                elif tag == "vPIS":
                    try:
                        impostos["pis"] = float(elem.text or 0)
                    except:
                        pass
                elif tag == "vCOFINS":
                    try:
                        impostos["cofins"] = float(elem.text or 0)
                    except:
                        pass
        except:
            pass
        
        return numero_nf, impostos
    
    # Processar dados
    dados = []
    for row in rows:
        row_dict = {k: row[k] for k in row.keys()}
        
        # Extrair dados do XML
        xml_str = row_dict.get("xml", "")
        numero_nf, impostos = extrair_dados_xml(xml_str)
        
        # Parse endereço JSON
        endereco_str = row_dict.get("endereco_entrega", "")
        endereco = {}
        if endereco_str:
            try:
                endereco = json.loads(endereco_str)
            except:
                endereco = {"endereco_raw": endereco_str}
        
        # Status e motivo
        status = row_dict.get("status", "pendente")
        if status == "identificado":
            status_texto = "Identificada"
            motivo = ""
        else:
            status_texto = "Pendente"
            motivo = row_dict.get("motivo_pendencia") or "Não identificado automaticamente"
        
        dados.append({
            "Chave NFe": row_dict.get("chave_acesso", ""),
            "Número NFe": numero_nf,
            "NSU": row_dict.get("nsu", ""),
            "Data Emissão": row_dict.get("data_emissao", ""),
            "Fornecedor": row_dict.get("fornecedor", ""),
            "CNPJ Fornecedor": row_dict.get("cnpj_emitente", ""),
            "Valor Total": row_dict.get("valor_total", 0),
            "ICMS": impostos.get("icms", 0),
            "IPI": impostos.get("ipi", 0),
            "PIS": impostos.get("pis", 0),
            "COFINS": impostos.get("cofins", 0),
            "Total Impostos": sum(impostos.values()),
            "Destinatário Nome": row_dict.get("nome_destinatario") or endereco.get("nome", ""),
            "Destinatário CNPJ": row_dict.get("cnpj_destinatario") or endereco.get("cnpj", ""),
            "Destinatário Endereço": endereco.get("endereco", ""),
            "Destinatário Cidade": endereco.get("cidade", ""),
            "Destinatário UF": endereco.get("uf", ""),
            "Destinatário CEP": endereco.get("cep", ""),
            "Informações Complementares": (row_dict.get("info_adicional", "") or "")[:500],
            "Status": status_texto,
            "Cliente Identificado": row_dict.get("cliente_identificado", ""),
            "Posto Identificado": row_dict.get("posto_identificado", ""),
            "Motivo Pendência": motivo
        })
    
    # Criar DataFrame
    df = pd.DataFrame(dados)
    
    # Gerar caminho do arquivo
    output_dir = Path(__file__).parent / "output"
    output_dir.mkdir(parents=True, exist_ok=True)
    
    timestamp = dt.now().strftime("%Y%m%d_%H%M%S")
    output_path = output_dir / f"nfes_exportacao_{timestamp}.xlsx"
    
    # Exportar para Excel
    with pd.ExcelWriter(str(output_path), engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="NFes", index=False)
    
    # Formatar Excel
    wb = load_workbook(str(output_path))
    ws = wb.active
    
    # Formatar cabeçalho
    header_fill = PatternFill("solid", fgColor="1a365d")
    header_font = Font(color="FFFFFF", bold=True)
    
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    
    # Ajustar largura de algumas colunas
    ws.column_dimensions["A"].width = 50  # Chave
    ws.column_dimensions["E"].width = 40  # Fornecedor
    ws.column_dimensions["S"].width = 60  # Info Complementares
    
    # Congelar primeira linha
    ws.freeze_panes = "A2"
    
    wb.save(str(output_path))
    
    print(f"[EXPORTACAO] Arquivo gerado: {output_path}")
    print(f"[EXPORTACAO] Total de linhas: {len(dados)}")
    
    return str(output_path)
